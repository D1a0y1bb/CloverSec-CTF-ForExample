import json
import sys
import tempfile
import unittest
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - CI may not install optional xlsx styling dependency.
    load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "plugins" / "cloversec-ctf-forexample" / "scripts"
sys.path.insert(0, str(SCRIPTS))

import cloversec_ctf_collect as collect
import cloversec_ctf_naming as naming
import cloversec_ctf_search as search


class CollectionHumanOutputTests(unittest.TestCase):
    @unittest.skipUnless(load_workbook, "openpyxl is optional; skip detailed xlsx style inspection when unavailable")
    def test_collection_human_xlsx_matches_golden_style(self):
        case = {
            "case_id": "case-1",
            "metadata": {
                "年份": "2024",
                "赛事来源": "02 imaginary ctf 2024",
                "赛事类型": "GitHub公开WP归档",
                "分类": "Web",
                "名称": "journal",
                "材料状态": "公开 WP + 附件/源码线索",
            },
            "asset_collection": {
                "writeup_candidates": [{"source_url": "https://github.com/demo/README.md"}],
                "attachment_candidates": [{"source_url": "https://github.com/demo/tree/main/journal"}],
            },
            "research": {},
            "evidence": [],
            "confidence": "medium",
        }
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "赛事题目信息收集表.xlsx"
            collect.write_collection_human_xlsx([case], output)
            workbook = load_workbook(output)
            sheet = workbook.active

        self.assertEqual(sheet.title, "Sheet1")
        self.assertIsNone(sheet.freeze_panes)
        self.assertEqual([sheet.cell(1, index).value for index in range(1, 11)], collect.COLLECTION_HUMAN_FIELDS)
        self.assertEqual(sheet["A2"].value, "2024")
        self.assertEqual(sheet["F2"].value, search.MATERIAL_STATUS_WP_PLUS_ASSET)
        self.assertEqual(sheet["I2"].value, None)
        self.assertEqual(sheet["J2"].value, None)
        self.assertAlmostEqual(sheet.column_dimensions["A"].width, 26.9326923076923, places=1)
        self.assertAlmostEqual(sheet.column_dimensions["J"].width, 13.0, places=1)
        self.assertAlmostEqual(sheet.row_dimensions[1].height, 18.05)
        self.assertAlmostEqual(sheet.row_dimensions[2].height, 18.05)
        for cell_ref, bold in [("A1", True), ("A2", False), ("J1", True), ("J2", False)]:
            cell = sheet[cell_ref]
            self.assertEqual(cell.font.name, "等线")
            self.assertEqual(cell.font.size, 12)
            self.assertEqual(cell.font.bold, bold)
            self.assertEqual(cell.alignment.horizontal, "center")
            self.assertEqual(cell.alignment.vertical, "center")
            self.assertIn(cell.alignment.wrap_text, (None, False))
            self.assertEqual(cell.fill.fill_type, "solid")
            self.assertEqual(cell.border.left.style, "thin")
            self.assertEqual(cell.border.right.style, "thin")
            self.assertEqual(cell.border.top.style, "thin")
            self.assertEqual(cell.border.bottom.style, "thin")

    def test_collection_human_row_normalizes_legacy_material_statuses(self):
        for value in ["未开始", "收集中", "已收集", "缺材料", "无法确认", "乱写"]:
            row = collect.collection_human_row({"metadata": {"材料状态": value}, "research": {}, "asset_collection": {}, "evidence": []})
            self.assertIn(row["已知材料情况"], search.COLLECTION_MATERIAL_STATUSES)

    def test_collection_machine_jsonl_keeps_case_id_and_machine_fields(self):
        case = {
            "case_id": "case-1",
            "metadata": {"名称": "baby sql", "分类": "Web", "材料状态": search.MATERIAL_STATUS_ASSET_ONLY},
            "research": {"gate": "can_continue"},
            "asset_collection": {},
            "evidence": [{"source_url": "https://example.com"}],
            "confidence": "high",
        }
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "collection_machine.jsonl"
            collect.write_collection_machine_jsonl([case], output)
            row = json.loads(output.read_text(encoding="utf-8").splitlines()[0])

        self.assertEqual(row["case_id"], "case-1")
        self.assertEqual(row["xlsx_fields"]["名称"], "baby sql")
        self.assertEqual(row["research"]["gate"], "can_continue")
        self.assertEqual(row["evidence"][0]["source_url"], "https://example.com")

    def test_naming_helpers_follow_archive_sample(self):
        metadata = {"分类": "Web", "名称": "EZSmuggler", "题目类型": "环境型"}

        self.assertEqual(naming.challenge_dir_name(metadata), "Web-EZSmuggler")
        self.assertEqual(naming.manual_filename({"题目分类": "Web", "题目标题": "EZsmuggler"}), "WEB-EZsmuggler.md")
        self.assertEqual(naming.challenge_subdirs(metadata), ["题目源码", "题目镜像", "题目手册"])
        self.assertEqual(naming.challenge_subdirs("附件型"), ["题目附件", "题目手册"])


if __name__ == "__main__":
    unittest.main()
