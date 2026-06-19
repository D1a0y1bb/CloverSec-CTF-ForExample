#!/usr/bin/env python3
"""CloverSec CTF data model helpers.

The default reader/writer works without third-party packages. When openpyxl is
available, xlsx export uses it for a cleaner table style and falls back to the
standard-library writer when it is not installed.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import cloversec_ctf_search as search


XLSX_FIELDS = [
    "HUB编号",
    "赛事来源",
    "题目来源",
    "名称",
    "分类",
    "题目类型",
    "Flag类型",
    "Flag",
    "难度编号",
    "星级",
    "分值",
    "资源等级",
    "提交时间",
    "开放端口",
    "离线/在线解题",
    "解题工具",
    "提交用户名",
    "审核人",
    "是否通过",
    "问题",
    "是否归档",
    "材料状态",
    "构建状态",
    "手册状态",
    "验证状态",
    "归档目录",
    "环境包/附件包路径",
    "备注",
]

REQUIRED_XLSX_FIELDS = ["名称", "分类", "题目类型", "Flag类型", "Flag", "验证状态"]

FLAG_TYPES = {"static", "dynamic", "unknown", "静态Flag", "动态Flag", "未知", "静态", "动态"}
FLAG_TYPE_DISPLAY_MAP = {
    "static": "静态Flag",
    "dynamic": "动态Flag",
    "unknown": "未知",
    "静态": "静态Flag",
    "动态": "动态Flag",
    "静态flag": "静态Flag",
    "动态flag": "动态Flag",
    "静态Flag": "静态Flag",
    "动态Flag": "动态Flag",
    "未知": "未知",
}
CONFIDENCE_VALUES = {"high", "medium", "low"}

STATUS_ENUMS = {
    "材料状态": set(search.COLLECTION_MATERIAL_STATUSES),
    "构建状态": {"不适用", "未开始", "已构建", "验证失败"},
    "手册状态": {"未开始", "草稿", "正式", "待人工完善", "已校验"},
    "验证状态": {"未验证", "部分通过", "通过", "失败"},
    "是否归档": {"是", "否"},
    "是否通过": {"是", "否"},
}

LIST_FIELDS = ["source_files", "attachments", "evidence"]
DICT_FIELDS = [
    "metadata",
    "research",
    "asset_collection",
    "flag",
    "environment",
    "docker_artifacts",
    "writeup",
    "hub_fields",
    "review",
    "archive",
]

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PACKAGE_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def validate_case(case: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if not isinstance(case, dict):
        return ["case must be a JSON object"]

    if not _text(case.get("case_id")):
        errors.append("case_id is required")

    for field in DICT_FIELDS:
        if field in case and not isinstance(case[field], dict):
            errors.append(f"{field} must be an object")

    for field in LIST_FIELDS:
        if field in case and not isinstance(case[field], list):
            errors.append(f"{field} must be an array")

    metadata = case.get("metadata") if isinstance(case.get("metadata"), dict) else {}
    flag = case.get("flag") if isinstance(case.get("flag"), dict) else {}

    if not _text(flag.get("value")):
        errors.append("flag.value is required and must contain the full Flag")
    if _text(flag.get("type")) and str(flag.get("type")) not in FLAG_TYPES:
        errors.append(f"flag.type has unsupported value: {flag.get('type')}")
    if flag.get("sensitive") is not True:
        errors.append("flag.sensitive must be true")

    confidence = case.get("confidence")
    if _text(confidence) and str(confidence) not in CONFIDENCE_VALUES:
        errors.append(f"confidence has unsupported value: {confidence}")

    row = case_to_xlsx_row(case)
    for field in REQUIRED_XLSX_FIELDS:
        if not _text(row.get(field)):
            errors.append(f"{field} is required for xlsx output")

    for field, allowed in STATUS_ENUMS.items():
        value = metadata.get(field)
        if _text(value) and str(value) not in allowed:
            errors.append(f"{field} has unsupported value: {value}")

    return errors


def validate_rows(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    for index, row in enumerate(rows, start=2):
        for field in REQUIRED_XLSX_FIELDS:
            if not _text(row.get(field)):
                errors.append(f"row {index}: {field} is required")
        for field, allowed in STATUS_ENUMS.items():
            value = row.get(field)
            if _text(value) and value not in allowed:
                errors.append(f"row {index}: {field} has unsupported value: {value}")
    return errors


def case_to_xlsx_row(case: dict[str, Any]) -> dict[str, str]:
    metadata = case.get("metadata") if isinstance(case.get("metadata"), dict) else {}
    flag = case.get("flag") if isinstance(case.get("flag"), dict) else {}
    row = {field: _string(metadata.get(field, "")) for field in XLSX_FIELDS}
    row["Flag"] = _string(flag.get("value") or metadata.get("Flag") or row["Flag"])
    row["Flag类型"] = normalize_flag_type_for_display(metadata.get("Flag类型") or flag.get("type") or row["Flag类型"])
    return row


def normalize_flag_type_for_display(value: Any) -> str:
    text = _string(value).strip()
    if not text:
        return ""
    return FLAG_TYPE_DISPLAY_MAP.get(text, FLAG_TYPE_DISPLAY_MAP.get(text.lower(), text))


def normalize_material_status(value: Any = "", *, has_writeup: bool = False, has_asset: bool = False, has_official: bool = False) -> str:
    return search.normalize_material_status(value, has_writeup=has_writeup, has_asset=has_asset, has_official=has_official)


def load_cases(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path)
    text = source.read_text(encoding="utf-8")
    if source.suffix.lower() == ".jsonl":
        cases = []
        for line_number, line in enumerate(text.splitlines(), start=1):
            if not line.strip():
                continue
            try:
                value = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{source}:{line_number}: invalid JSON: {exc}") from exc
            if not isinstance(value, dict):
                raise ValueError(f"{source}:{line_number}: JSONL item must be an object")
            cases.append(value)
        return cases

    value = json.loads(text)
    if isinstance(value, list):
        if not all(isinstance(item, dict) for item in value):
            raise ValueError(f"{source}: all JSON array items must be objects")
        return value
    if isinstance(value, dict):
        return [value]
    raise ValueError(f"{source}: expected object, array, or JSONL")


def write_xlsx(cases: list[dict[str, Any]], path: str | Path) -> None:
    rows = [case_to_xlsx_row(case) for case in cases]
    write_rows_xlsx(rows, path)


def write_rows_xlsx(rows: list[dict[str, Any]], path: str | Path) -> None:
    write_table_xlsx(rows, path, fields=XLSX_FIELDS, sheet_name="归档表")


def write_table_xlsx(rows: list[dict[str, Any]], path: str | Path, *, fields: list[str] | None = None, sheet_name: str = "归档表") -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    headers = fields or infer_table_fields(rows)
    if os_environ_flag("CLOVERSEC_XLSX_WRITER") != "stdlib" and write_table_xlsx_openpyxl(rows, output, headers=headers, sheet_name=sheet_name):
        return
    matrix = [headers]
    for row in rows:
        matrix.append([_string(row.get(field, "")) for field in headers])

    sheet_xml = _worksheet_xml(matrix)
    files = {
        "[Content_Types].xml": _content_types_xml(),
        "_rels/.rels": _root_rels_xml(),
        "xl/workbook.xml": _workbook_xml(sheet_name=sheet_name),
        "xl/_rels/workbook.xml.rels": _workbook_rels_xml(),
        "xl/worksheets/sheet1.xml": sheet_xml,
        "docProps/core.xml": _core_xml(),
        "docProps/app.xml": _app_xml(),
    }
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)


def write_styled_xlsx(
    rows: list[dict[str, Any]],
    path: str | Path,
    *,
    fields: list[str],
    sheet_name: str,
    style: dict[str, Any],
) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    matrix = [fields]
    for row in rows:
        matrix.append([_string(row.get(field, "")) for field in fields])
    column_widths = [float(style.get("column_widths", {}).get(field, suggested_column_width(field))) for field in fields]
    sheet_xml = _worksheet_xml(
        matrix,
        column_widths=column_widths,
        row_height=float(style.get("row_height", 18.05)),
        header_style=1,
        body_style=2,
    )
    files = {
        "[Content_Types].xml": _content_types_xml(include_styles=True),
        "_rels/.rels": _root_rels_xml(),
        "xl/workbook.xml": _workbook_xml(sheet_name=sheet_name),
        "xl/_rels/workbook.xml.rels": _workbook_rels_xml(include_styles=True),
        "xl/worksheets/sheet1.xml": sheet_xml,
        "xl/styles.xml": _styles_xml(
            font_name=str(style.get("font_name") or "等线"),
            font_size=float(style.get("font_size") or 12),
        ),
        "docProps/core.xml": _core_xml(),
        "docProps/app.xml": _app_xml(),
    }
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)


def write_table_xlsx_openpyxl(rows: list[dict[str, Any]], path: Path, *, headers: list[str], sheet_name: str) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return False

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31] or "归档表"
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="等线", size=12, bold=True)
    body_font = Font(name="等线", size=12)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.append(headers)
    for row in rows:
        sheet.append([_string(row.get(field, "")) for field in headers])

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            cell.font = header_font if cell.row == 1 else body_font
    for index, field in enumerate(headers, start=1):
        width = suggested_column_width(field)
        for row in rows[:200]:
            width = max(width, min(len(_string(row.get(field, ""))) + 2, 42))
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 24
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return True


def suggested_column_width(field: str) -> int:
    widths = {
        "名称": 24,
        "赛事来源": 22,
        "题目来源": 22,
        "分类": 12,
        "Flag": 28,
        "问题": 36,
        "备注": 34,
        "归档目录": 34,
        "环境包/附件包路径": 34,
    }
    return widths.get(field, max(12, min(len(field) + 4, 24)))


def os_environ_flag(name: str) -> str:
    try:
        import os
    except Exception:  # noqa: BLE001
        return ""
    return os.environ.get(name, "").strip().lower()


def infer_table_fields(rows: list[dict[str, Any]]) -> list[str]:
    fields: list[str] = []
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(str(key))
    return fields or ["结果"]


def read_xlsx(path: str | Path) -> list[dict[str, str]]:
    matrix = _read_xlsx_matrix(Path(path))
    return matrix_to_rows(matrix)


def read_xlsx_sheets(path: str | Path) -> dict[str, list[dict[str, str]]]:
    source = Path(path)
    with zipfile.ZipFile(source) as archive:
        shared_strings = _read_shared_strings(archive)
        sheet_paths = _sheet_paths(archive)
        sheets = {}
        for sheet_name, sheet_path in sheet_paths:
            root = ET.fromstring(archive.read(sheet_path))
            matrix = _worksheet_matrix(root, shared_strings)
            sheets[sheet_name] = matrix_to_rows(matrix)
        return sheets


def matrix_to_rows(matrix: list[list[str]]) -> list[dict[str, str]]:
    if not matrix:
        return []
    headers = [cell.strip() for cell in matrix[0]]
    rows: list[dict[str, str]] = []
    for raw_row in matrix[1:]:
        if not any(_text(cell) for cell in raw_row):
            continue
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row[header] = raw_row[index] if index < len(raw_row) else ""
        for field in XLSX_FIELDS:
            row.setdefault(field, "")
        rows.append({field: row.get(field, "") for field in XLSX_FIELDS})
    return rows


def summarize_cases(cases: list[dict[str, Any]]) -> dict[str, Any]:
    rows = [case_to_xlsx_row(case) for case in cases]
    missing = {
        field: sum(1 for row in rows if not _text(row.get(field)))
        for field in REQUIRED_XLSX_FIELDS
    }
    validation_status = Counter(row.get("验证状态", "") or "未填写" for row in rows)
    archive_status = Counter(row.get("是否归档", "") or "未填写" for row in rows)
    return {
        "total": len(cases),
        "by_validation_status": dict(validation_status),
        "by_archive_status": dict(archive_status),
        "missing_required_fields": missing,
    }


def _read_xlsx_matrix(path: Path) -> list[list[str]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = _read_shared_strings(archive)
        sheet_name = _first_sheet_path(archive)
        root = ET.fromstring(archive.read(sheet_name))

    return _worksheet_matrix(root, shared_strings)


def _worksheet_matrix(root: ET.Element, shared_strings: list[str]) -> list[list[str]]:
    rows: list[list[str]] = []
    for row_node in root.findall(f".//{{{NS_MAIN}}}sheetData/{{{NS_MAIN}}}row"):
        values: dict[int, str] = {}
        max_col = 0
        for cell in row_node.findall(f"{{{NS_MAIN}}}c"):
            ref = cell.attrib.get("r", "")
            col = _column_index(ref)
            if col < 1:
                col = max_col + 1
            max_col = max(max_col, col)
            values[col] = _cell_value(cell, shared_strings)
        rows.append([values.get(col, "") for col in range(1, max_col + 1)])
    return rows


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    values = []
    for item in root.findall(f"{{{NS_MAIN}}}si"):
        texts = [node.text or "" for node in item.findall(f".//{{{NS_MAIN}}}t")]
        values.append("".join(texts))
    return values


def _first_sheet_path(archive: zipfile.ZipFile) -> str:
    sheet_paths = _sheet_paths(archive)
    if not sheet_paths:
        raise ValueError("workbook has no sheets")
    return sheet_paths[0][1]


def _sheet_paths(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {
        rel.attrib.get("Id"): rel.attrib.get("Target", "")
        for rel in rels.findall(f"{{{NS_PACKAGE_REL}}}Relationship")
    }
    sheet_paths: list[tuple[str, str]] = []
    for sheet in workbook.findall(f".//{{{NS_MAIN}}}sheet"):
        rel_id = sheet.attrib.get(f"{{{NS_REL}}}id")
        target = targets.get(rel_id, "")
        if not target:
            continue
        normalized = target.lstrip("/")
        if not normalized.startswith("xl/"):
            normalized = "xl/" + normalized
        sheet_paths.append((sheet.attrib.get("name", ""), normalized))
    return sheet_paths


def _cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        texts = [node.text or "" for node in cell.findall(f".//{{{NS_MAIN}}}t")]
        return "".join(texts)
    value_node = cell.find(f"{{{NS_MAIN}}}v")
    value = value_node.text if value_node is not None and value_node.text is not None else ""
    if cell_type == "s" and value:
        index = int(value)
        return shared_strings[index] if index < len(shared_strings) else ""
    return value


def _worksheet_xml(
    matrix: list[list[str]],
    *,
    column_widths: list[float] | None = None,
    row_height: float | None = None,
    header_style: int | None = None,
    body_style: int | None = None,
) -> str:
    row_xml = []
    for row_index, row in enumerate(matrix, start=1):
        cell_xml = []
        for col_index, value in enumerate(row, start=1):
            ref = f"{_column_name(col_index)}{row_index}"
            style_id = header_style if row_index == 1 else body_style
            style_attr = f' s="{style_id}"' if style_id is not None else ""
            if _string(value) == "":
                cell_xml.append(f'<c r="{ref}"{style_attr}/>')
                continue
            cell_xml.append(
                f'<c r="{ref}"{style_attr} t="inlineStr"><is><t xml:space="preserve">'
                f"{_xml_escape(value)}</t></is></c>"
            )
        height_attr = f' ht="{row_height:g}" customHeight="1"' if row_height is not None else ""
        row_xml.append(f'<row r="{row_index}"{height_attr}>{"".join(cell_xml)}</row>')
    last_ref = f"{_column_name(len(matrix[0]))}{len(matrix)}" if matrix else "A1"
    cols_xml = ""
    if column_widths:
        cols = []
        for index, width in enumerate(column_widths, start=1):
            cols.append(f'<col min="{index}" max="{index}" width="{width:g}" customWidth="1"/>')
        cols_xml = f"<cols>{''.join(cols)}</cols>"
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet xmlns="{NS_MAIN}" xmlns:r="{NS_REL}">'
        f'<dimension ref="A1:{last_ref}"/>'
        f"{cols_xml}"
        f'<sheetData>{"".join(row_xml)}</sheetData>'
        "</worksheet>"
    )


def _content_types_xml(*, include_styles: bool = False) -> str:
    styles = (
        '  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>\n'
        if include_styles
        else ""
    )
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
{styles}  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>"""


def _root_rels_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>"""


def _workbook_xml(*, sheet_name: str = "归档表") -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<workbook xmlns="{NS_MAIN}" xmlns:r="{NS_REL}">'
        f'<sheets><sheet name="{_xml_escape(sheet_name)}" sheetId="1" r:id="rId1"/></sheets>'
        "</workbook>"
    )


def _workbook_rels_xml(*, include_styles: bool = False) -> str:
    styles_rel = (
        '  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>\n'
        if include_styles
        else ""
    )
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
{styles_rel}</Relationships>"""


def _styles_xml(*, font_name: str, font_size: float) -> str:
    size = f"{font_size:g}"
    escaped_font = _xml_escape(font_name)
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="{NS_MAIN}">
  <fonts count="3">
    <font><sz val="11"/><name val="Calibri"/></font>
    <font><b/><sz val="{size}"/><name val="{escaped_font}"/></font>
    <font><sz val="{size}"/><name val="{escaped_font}"/></font>
  </fonts>
  <fills count="3">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFFFFFF"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="2">
    <border><left/><right/><top/><bottom/><diagonal/></border>
    <border>
      <left style="thin"><color rgb="FFD9D9D9"/></left>
      <right style="thin"><color rgb="FFD9D9D9"/></right>
      <top style="thin"><color rgb="FFD9D9D9"/></top>
      <bottom style="thin"><color rgb="FFD9D9D9"/></bottom>
      <diagonal/>
    </border>
  </borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="3">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>
    <xf numFmtId="0" fontId="2" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
  <dxfs count="0"/>
  <tableStyles count="0" defaultTableStyle="TableStyleMedium2" defaultPivotStyle="PivotStyleLight16"/>
</styleSheet>"""


def _core_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
 xmlns:dc="http://purl.org/dc/elements/1.1/">
  <dc:creator>CloverSec CTF For Example</dc:creator>
</cp:coreProperties>"""


def _app_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
 xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>CloverSec CTF For Example</Application>
</Properties>"""


def _column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _column_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref.upper())
    if not match:
        return 0
    value = 0
    for char in match.group(1):
        value = value * 26 + ord(char) - 64
    return value


def _xml_escape(value: Any) -> str:
    return (
        _string(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _string(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _text(value: Any) -> bool:
    return bool(_string(value).strip())


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="CloverSec CTF data model utilities")
    subparsers = parser.add_subparsers(dest="command", required=True)

    validate_parser = subparsers.add_parser("validate-json", help="validate ctf_case JSON or JSONL")
    validate_parser.add_argument("input")

    export_parser = subparsers.add_parser("export-xlsx", help="export JSON/JSONL cases to xlsx")
    export_parser.add_argument("input")
    export_parser.add_argument("output")

    import_parser = subparsers.add_parser("import-xlsx", help="read xlsx and emit JSON rows")
    import_parser.add_argument("input")
    import_parser.add_argument("--output-jsonl")

    summary_parser = subparsers.add_parser("summary", help="summarize JSON/JSONL cases")
    summary_parser.add_argument("input")

    args = parser.parse_args(argv)

    if args.command == "validate-json":
        cases = load_cases(args.input)
        errors = []
        for index, case in enumerate(cases, start=1):
            for error in validate_case(case):
                errors.append(f"case {index}: {error}")
        if errors:
            print("\n".join(errors), file=sys.stderr)
            return 1
        print(f"validated {len(cases)} case(s)")
        return 0

    if args.command == "export-xlsx":
        cases = load_cases(args.input)
        all_errors = []
        for index, case in enumerate(cases, start=1):
            for error in validate_case(case):
                all_errors.append(f"case {index}: {error}")
        if all_errors:
            print("\n".join(all_errors), file=sys.stderr)
            return 1
        write_xlsx(cases, args.output)
        print(f"wrote {args.output}")
        return 0

    if args.command == "import-xlsx":
        rows = read_xlsx(args.input)
        errors = validate_rows(rows)
        if errors:
            print("\n".join(errors), file=sys.stderr)
            return 1
        if args.output_jsonl:
            output = Path(args.output_jsonl)
            output.write_text(
                "\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + "\n",
                encoding="utf-8",
            )
            print(f"wrote {output}")
        else:
            print(json.dumps(rows, ensure_ascii=False, indent=2))
        return 0

    if args.command == "summary":
        cases = load_cases(args.input)
        print(json.dumps(summarize_cases(cases), ensure_ascii=False, indent=2))
        return 0

    parser.error("unknown command")
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
